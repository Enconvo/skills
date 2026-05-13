#!/usr/bin/env python3
"""
Portfolio XLSX Report Generator
Runs portfolio_checker.py --json, then populates the STYLE-02 Executive Editorial template.
Usage: python3 generate_portfolio_xlsx.py [--output path.xlsx]

Template: templates/Portfolio_Report_Template.xlsx (STYLE-02, tweaked by user)
Data source: portfolio_checker.py --json
"""
import os, sys, json, subprocess, argparse
from datetime import date, datetime
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, SCRIPT_DIR)
from config import get_output_dir  # noqa: E402

TEMPLATE_PATH = os.path.join(SCRIPT_DIR, "templates", "Portfolio_Report_Template.xlsx")
DEFAULT_OUTPUT = str(get_output_dir() / f"Portfolio_Report_{date.today().isoformat()}.xlsx")

# ─── STYLE-02 Executive Editorial constants ───
PAL = {
    "heading": "C8102E", "body": "333333", "muted": "8C8279",
    "accent": "C8102E", "header_fill": "C8102E", "header_font": "FFFFFF",
    "alt_row_fill": "FAF7F2", "positive": "16A34A", "negative": "DC2626", "warning": "D97706",
}

# Account display names (redacted)
ACCOUNT_DISPLAY = {}
def fmt_code(code): return code
FONT_BODY = Font(name="Georgia", size=11, color=PAL["body"])
FONT_BODY_B = Font(name="Georgia", size=11, bold=True, color=PAL["body"])
FONT_POS = Font(name="Georgia", size=11, bold=True, color=PAL["positive"])
FONT_NEG = Font(name="Georgia", size=11, bold=True, color=PAL["negative"])
FONT_WARN = Font(name="Georgia", size=11, bold=True, color=PAL["warning"])
FONT_HDR = Font(name="Georgia", size=11, bold=True, color=PAL["header_font"])
FONT_METRIC_LG = Font(name="Georgia", size=20, bold=True)
FONT_METRIC_CAP = Font(name="Georgia", size=10, color="999999")
FONT_COVER_DATE = Font(name="Georgia", size=12, color="999999")

FILL_HDR = PatternFill("solid", fgColor=PAL["header_fill"])
FILL_ALT = PatternFill("solid", fgColor=PAL["alt_row_fill"])
FILL_POS_BG = PatternFill("solid", fgColor="ECFDF5")
FILL_NEG_BG = PatternFill("solid", fgColor="FEF2F2")
FILL_WARN_BG = PatternFill("solid", fgColor="FFFBEB")

BORDER_THIN = Border(bottom=Side(style="thin", color="E0D8CF"))
ALIGN_L = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_C = Alignment(horizontal="center", vertical="center")
ALIGN_R = Alignment(horizontal="right", vertical="center")


def get_status_style(status):
    """Return (font, fill) based on status string."""
    s = status.upper()
    if s in ("DANGER", "FAR"):
        return FONT_NEG, FILL_NEG_BG
    elif s in ("WARNING", "POSSIBLE", "CLOSE"):
        return FONT_WARN, FILL_WARN_BG
    else:
        return FONT_POS, FILL_POS_BG


def get_buffer_font(buf_pct):
    """Return font based on buffer percentage."""
    if buf_pct < 0:
        return FONT_NEG
    elif buf_pct < 20:
        return FONT_WARN
    return FONT_POS


def write_row(ws, row, col_start, values, fonts=None, fills=None, aligns=None):
    """Write a row of values with optional per-cell styling."""
    for i, val in enumerate(values):
        c = ws.cell(row=row, column=col_start + i, value=val)
        c.font = fonts[i] if fonts else FONT_BODY
        c.fill = fills[i] if fills else PatternFill()
        c.alignment = aligns[i] if aligns else ALIGN_L
        c.border = BORDER_THIN


def run_portfolio_checker():
    """Run portfolio_checker.py --json and return parsed data."""
    result = subprocess.run(
        [sys.executable, os.path.join(SCRIPT_DIR, "portfolio_checker.py"), "--json"],
        capture_output=True, text=True, cwd=SCRIPT_DIR
    )
    if result.returncode != 0:
        print(f"ERROR: portfolio_checker.py failed:\n{result.stderr}", file=sys.stderr)
        sys.exit(1)
    return json.loads(result.stdout)


def populate_cover(ws, data):
    """Update cover sheet KPIs and date."""
    today_str = datetime.now().strftime("%B %d, %Y")
    
    # Update date (B16)
    ws["B16"] = f"{today_str}  |  Prepared by Vivienne, Finance Specialist"
    ws["B16"].font = FONT_COVER_DATE
    
    # KPI values
    structured_notes = data.get("structured_notes") or data.get("notes", [])
    active = [n for n in structured_notes if n.get("status", "").upper() != "AUTOCALLED"]
    total_exposure = sum(n.get("notional", n.get("amount", 0)) for n in active)
    danger_count = sum(1 for n in active if n.get("status", "").upper() == "DANGER")
    
    autocall_ready = [n for n in data.get("autocall_pipeline") or data.get("autocall_proximity", []) 
                      if n.get("readiness", "").upper() == "READY" and n.get("days_to_obs", 999) <= 7]
    autocall_7d_amount = sum(n.get("notional", n.get("amount", 0)) for n in autocall_ready)
    
    ws["B20"] = f"${total_exposure/1e6:.1f}M"
    ws["D20"] = f"${autocall_7d_amount/1e3:.0f}K" if autocall_7d_amount > 0 else "$0"
    ws["F20"] = f"{danger_count} DANGER" if danger_count > 0 else "ALL SAFE"
    
    # Color the danger KPI
    if danger_count > 0:
        ws["F20"].font = Font(name="Georgia", size=20, bold=True, color=PAL["negative"])
    else:
        ws["F20"].font = Font(name="Georgia", size=20, bold=True, color=PAL["positive"])


def populate_dashboard(ws, data):
    """Update dashboard KPIs, risk summary, and blockers."""
    structured_notes = data.get("structured_notes") or data.get("notes", [])
    active = [n for n in structured_notes if n.get("status", "").upper() != "AUTOCALLED"]
    autocalled = [n for n in structured_notes if n.get("status", "").upper() == "AUTOCALLED"]
    
    total_exposure = sum(n.get("notional", n.get("amount", 0)) for n in active)
    danger = [n for n in active if n.get("status", "").upper() == "DANGER"]
    warnings = [n for n in active if n.get("status", "").upper() in ("WARNING", "CLOSE")]
    safe = [n for n in active if n.get("status", "").upper() == "SAFE"]
    
    autocall_ready = [n for n in data.get("autocall_pipeline") or data.get("autocall_proximity", [])
                      if n.get("readiness", "").upper() == "READY" and n.get("days_to_obs", 999) <= 7]
    autocall_7d_amount = sum(n.get("notional", n.get("amount", 0)) for n in autocall_ready)
    
    # KPI row (row 4)
    kpi_vals = [
        (str(len(active)), PAL["heading"]),
        (f"${total_exposure/1e6:.1f}M", PAL["heading"]),
        (f"${autocall_7d_amount/1e3:.0f}K", PAL["positive"]),
        (str(len(danger)), PAL["negative"]),
        (str(len(warnings)), PAL["warning"]),
        (str(len(autocalled)), PAL["positive"]),
    ]
    for i, (val, color) in enumerate(kpi_vals):
        c = ws.cell(row=4, column=2 + i, value=val)
        c.font = Font(name="Georgia", size=18, bold=True, color=color)
    
    # Risk summary (rows 11-13)
    risk_rows = [
        ("DANGER", len(danger), sum(n.get("notional", n.get("amount", 0)) for n in danger),
         danger[0].get("worst_performer", "N/A") if danger else "N/A"),
        ("WARNING", len(warnings), sum(n.get("notional", n.get("amount", 0)) for n in warnings),
         warnings[0].get("worst_performer", "N/A") if warnings else "N/A"),
        ("SAFE", len(safe), sum(n.get("notional", n.get("amount", 0)) for n in safe),
         "All comfortably above strikes"),
    ]
    for j, (status, count, exp, concern) in enumerate(risk_rows):
        r = 11 + j
        sfont, fill = get_status_style(status)
        ws.cell(row=r, column=2, value=status).font = sfont
        ws.cell(row=r, column=2).fill = fill; ws.cell(row=r, column=2).alignment = ALIGN_C
        ws.cell(row=r, column=4, value=str(count)).font = FONT_BODY
        ws.cell(row=r, column=3).fill = fill; ws.cell(row=r, column=3).alignment = ALIGN_C
        ws.cell(row=r, column=5, value=f"${exp:,.0f}").font = FONT_BODY_B
        ws.cell(row=r, column=4).fill = fill; ws.cell(row=r, column=4).alignment = ALIGN_R
        ws.cell(row=r, column=6, value=concern).font = FONT_BODY
        ws.cell(row=r, column=5).fill = fill; ws.cell(row=r, column=5).alignment = ALIGN_L
        for col in [2,3,4,5]:
            ws.cell(row=r, column=col).border = BORDER_THIN
    
    # Blockers (rows 19+) — clear stale rows first
    for rr in range(19, 25):
        for cc in [2, 3, 4, 5]:
            ws.cell(row=rr, column=cc, value="")
            ws.cell(row=rr, column=cc).fill = PatternFill()
            ws.cell(row=rr, column=cc).border = Border()

    blockers = data.get("blockers", [])
    if not blockers:
        ws.cell(row=19, column=2, value="N/A").font = FONT_BODY_B
        ws.cell(row=19, column=3, value="-").font = FONT_BODY
        ws.cell(row=19, column=4, value="-").font = FONT_BODY
        ws.cell(row=19, column=5, value="-").font = FONT_BODY
        for col in [2,3,4,5]:
            ws.cell(row=19, column=col).alignment = ALIGN_C
            ws.cell(row=19, column=col).fill = FILL_ALT
            ws.cell(row=19, column=col).border = BORDER_THIN
    else:
        for j, blk in enumerate(blockers[:5]):
            r = 19 + j
            ws.cell(row=r, column=2, value=blk.get("ticker", "")).font = FONT_BODY_B
            ws.cell(row=r, column=2).alignment = ALIGN_C; ws.cell(row=r, column=2).fill = FILL_ALT
            ws.cell(row=r, column=4, value=f"{blk.get('blocks', 0)} note(s)").font = FONT_BODY
            ws.cell(row=r, column=3).alignment = ALIGN_C; ws.cell(row=r, column=3).fill = FILL_ALT
            ws.cell(row=r, column=5, value=f"${blk.get('exposure', 0):,.0f}").font = FONT_BODY_B
            ws.cell(row=r, column=4).alignment = ALIGN_R; ws.cell(row=r, column=4).fill = FILL_ALT
            ws.cell(row=r, column=6, value=f"{blk.get('avg_gap', 0):+.1f}%").font = FONT_NEG
            ws.cell(row=r, column=5).alignment = ALIGN_C; ws.cell(row=r, column=5).fill = FILL_ALT
            for col in [2,3,4,5]:
                ws.cell(row=r, column=col).border = BORDER_THIN


def populate_structured_notes(ws, data):
    """Populate Structured Notes sheet — all underlyings with full price details."""
    structured_notes = data.get("structured_notes") or data.get("notes", [])
    active = [n for n in structured_notes if n.get("status", "").upper() != "AUTOCALLED"]
    autocalled = [n for n in structured_notes if n.get("status", "").upper() == "AUTOCALLED"]

    # Clear old data
    merges_to_remove = [m for m in ws.merged_cells.ranges if m.min_row >= 4]
    for m in merges_to_remove:
        ws.unmerge_cells(str(m))
    for r in range(4, 80):
        for c in range(2, 16):
            cell = ws.cell(row=r, column=c)
            cell.value = None; cell.font = FONT_BODY; cell.fill = PatternFill(); cell.border = Border()

    # Column widths
    widths = {'B': 9, 'C': 14, 'D': 10, 'E': 30, 'F': 12, 'G': 8, 'H': 12, 'I': 12,
              'J': 10, 'K': 10, 'L': 10, 'M': 10, 'N': 10, 'O': 10, 'P': 7}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Headers: B=Acct C=ISIN D=Status E=Description F=Notional G=Coupon H=Expiry I=Next Obs
    #          J=Underlying K=Current L=Strike M=Call/AC N=vs Strike O=vs Call P=Days
    headers = ["Code", "ISIN", "Status", "Description", "Notional", "Coupon", "Expiry", "Next Obs",
               "Underlying", "Current", "Strike", "Call/AC", "vs Strike", "vs Call", "Days"]
    for i, h in enumerate(headers):
        c = ws.cell(row=4, column=2 + i, value=h)
        c.font = FONT_HDR; c.fill = FILL_HDR; c.alignment = ALIGN_C; c.border = BORDER_THIN

    def sc(row, col, val, font=FONT_BODY, align=ALIGN_L, fill_=None):
        """Set cell helper."""
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = font; cell.alignment = align; cell.border = BORDER_THIN
        if fill_: cell.fill = fill_
        return cell

    r = 5
    for note in active:
        status = note.get("status", "SAFE").upper()
        sfont, nfill = get_status_style(status)
        underlyings = note.get("underlyings", [])
        n_ul = max(len(underlyings), 1)
        days = note.get("days_left", note.get("days_to_expiry", 0))

        # Note-level cells (first row only)
        sc(r, 2, fmt_code(note.get("account", "")), FONT_BODY_B, ALIGN_C, nfill)
        sc(r, 3, note.get("isin", ""), FONT_BODY, ALIGN_L, nfill)
        sc(r, 4, status, sfont, ALIGN_C, nfill)
        sc(r, 5, note.get("description") or note.get("product", ""), FONT_BODY, ALIGN_L, nfill)
        sc(r, 6, f"${note.get('notional', note.get('amount', 0)):,.0f}", FONT_BODY_B, ALIGN_R, nfill)
        sc(r, 7, note.get("coupon", ""), FONT_BODY, ALIGN_C, nfill)
        sc(r, 8, note.get("expiry", ""), FONT_BODY, ALIGN_C, nfill)
        # Next Observation date
        next_obs = note.get("next_observation", "")
        days_to_obs = note.get("days_to_next_obs")
        obs_label = f"{next_obs} ({days_to_obs}d)" if next_obs and days_to_obs is not None else (next_obs or "—")
        sc(r, 9, obs_label, FONT_BODY_B if days_to_obs and days_to_obs <= 7 else FONT_BODY, ALIGN_C, nfill)
        sc(r, 16, days, FONT_BODY, ALIGN_C, nfill)

        # Each underlying gets its own row
        for k, u in enumerate(underlyings):
            ur = r + k
            ticker = u.get("ticker", "")
            current = u.get("current", 0)
            strike = u.get("strike", 0)
            call = u.get("call_price", 0)
            vs_strike = ((current - strike) / strike * 100) if strike else 0
            vs_call = ((current - call) / call * 100) if call else 0
            is_worst = (ticker == note.get("worst_performer", ""))

            tfont = FONT_BODY_B if is_worst else FONT_BODY
            sc(ur, 10, ("⚠ " + ticker) if is_worst else ticker, tfont, ALIGN_L, nfill)
            sc(ur, 11, f"${current:,.2f}", FONT_BODY_B, ALIGN_R, nfill)
            sc(ur, 12, f"${strike:,.2f}", FONT_BODY, ALIGN_R, nfill)
            sc(ur, 13, f"${call:,.2f}", FONT_BODY, ALIGN_R, nfill)

            sf = FONT_NEG if vs_strike < 0 else (FONT_WARN if vs_strike < 20 else FONT_POS)
            sc(ur, 14, f"{vs_strike:+.1f}%", sf, ALIGN_C, nfill)

            cf = FONT_NEG if vs_call < 0 else FONT_POS
            sc(ur, 15, f"{vs_call:+.1f}%", cf, ALIGN_C, nfill)

            # Fill blank note-level cells on sub-rows
            if k > 0:
                for col in [2, 3, 4, 5, 6, 7, 8, 9, 16]:
                    cell = ws.cell(row=ur, column=col)
                    cell.fill = nfill; cell.border = BORDER_THIN

        r += n_ul

    # Autocalled section
    ar = r + 2
    ws.merge_cells(f"B{ar}:J{ar}")
    ws.cell(row=ar, column=2, value="Autocalled (Redeemed Early)").font = Font(name="Georgia", size=13, bold=True, color=PAL["body"])
    hr = ar + 2
    for i, h in enumerate(["ISIN", "Status", "Description", "Notional", "Coupon", "Autocall Date"]):
        c = ws.cell(row=hr, column=2 + i, value=h)
        c.font = FONT_HDR; c.fill = FILL_HDR; c.alignment = ALIGN_C
    for j, note in enumerate(autocalled):
        cr = hr + 1 + j
        sc(cr, 2, note.get("isin", ""), FONT_BODY, ALIGN_L, FILL_POS_BG)
        sc(cr, 3, "AUTOCALLED", FONT_POS, ALIGN_C, FILL_POS_BG)
        sc(cr, 4, note.get("description", ""), FONT_BODY, ALIGN_L, FILL_POS_BG)
        sc(cr, 5, f"${note.get('notional', 0):,.0f}", FONT_BODY_B, ALIGN_R, FILL_POS_BG)
        sc(cr, 6, note.get("coupon", ""), FONT_BODY, ALIGN_C, FILL_POS_BG)
        sc(cr, 7, note.get("autocall_date", ""), FONT_BODY, ALIGN_C, FILL_POS_BG)


def populate_sell_puts(ws, data):
    """Populate Sell Puts sheet with account info."""
    puts = data.get("sell_puts", [])

    merges_to_remove = [m for m in ws.merged_cells.ranges if m.min_row >= 4]
    for m in merges_to_remove:
        ws.unmerge_cells(str(m))
    for r in range(4, 20):
        for c in range(2, 10):
            cell = ws.cell(row=r, column=c)
            cell.value = None; cell.fill = PatternFill(); cell.border = Border()

    # Headers: B=Acct C=Ticker D=Strike E=Current F=Buffer G=Expiry H=Status
    headers = ["Code", "Ticker", "Strike", "Current", "Buffer", "Expiry", "Status"]
    for i, h in enumerate(headers):
        c = ws.cell(row=4, column=2 + i, value=h)
        c.font = FONT_HDR; c.fill = FILL_HDR; c.alignment = ALIGN_C; c.border = BORDER_THIN

    ws.column_dimensions['B'].width = 9
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 10

    def sc(row, col, val, font=FONT_BODY, align=ALIGN_L, fill_=None):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = font; cell.alignment = align; cell.border = BORDER_THIN
        if fill_: cell.fill = fill_

    active_puts = [p for p in puts if p.get("status", "").upper() != "CLOSED"]
    for j, p in enumerate(active_puts):
        r = 5 + j
        fill = FILL_ALT if j % 2 == 0 else PatternFill()
        buf = p.get("buffer_pct", 0)

        sc(r, 2, fmt_code(p.get("account", "")), FONT_BODY_B, ALIGN_C, fill)
        sc(r, 3, p.get("ticker", ""), FONT_BODY_B, ALIGN_L, fill)
        sc(r, 4, f"${p.get('strike', 0):.2f}", FONT_BODY, ALIGN_R, fill)
        sc(r, 5, f"${p.get('current', 0):.2f}", FONT_BODY, ALIGN_R, fill)
        sc(r, 6, f"{buf:+.1f}%", get_buffer_font(buf), ALIGN_C, fill)
        sc(r, 7, p.get("expiry", ""), FONT_BODY, ALIGN_C, fill)
        sc(r, 8, "SAFE", FONT_POS, ALIGN_C, fill)


def populate_fx_accumulators(ws, data):
    """Populate FX & Accumulators sheet with account info."""
    fx = data.get("fx_options", [])
    accum = data.get("accumulators", [])

    # FX Options header
    fx_headers = ["ID", "Description", "Notional", "Spot/Strike", "Buffer", "Expiry"]
    for i, h in enumerate(fx_headers):
        c = ws.cell(row=4, column=2 + i, value=h)
        c.font = FONT_HDR; c.fill = FILL_HDR; c.alignment = ALIGN_C; c.border = BORDER_THIN

    if fx:
        f = fx[0]
        def sc(row, col, val, font=FONT_BODY, align=ALIGN_L, fill_=None):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = font; cell.alignment = align; cell.border = BORDER_THIN
            if fill_: cell.fill = fill_

        sc(5, 2, f.get("id", ""), FONT_BODY, ALIGN_L, FILL_POS_BG)
        sc(5, 3, f.get("description", ""), FONT_BODY, ALIGN_L, FILL_POS_BG)
        sc(5, 4, f"${f.get('notional', 0):,.0f}", FONT_BODY_B, ALIGN_R, FILL_POS_BG)
        sc(5, 5, f.get("spot_strike", ""), FONT_BODY, ALIGN_C, FILL_POS_BG)
        sc(5, 6, f.get("buffer", ""), FONT_POS, ALIGN_C, FILL_POS_BG)
        sc(5, 7, f.get("expiry", ""), FONT_BODY, ALIGN_C, FILL_POS_BG)

    # Accumulators
    merges_to_remove = [m for m in ws.merged_cells.ranges if m.min_row >= 10]
    for m in merges_to_remove:
        ws.unmerge_cells(str(m))
    for r in range(10, 18):
        for c in range(2, 9):
            cell = ws.cell(row=r, column=c)
            cell.value = None; cell.fill = PatternFill(); cell.border = Border()

    # Accumulator headers
    acc_headers = ["Code", "Description", "Strike/Call", "Current", "Buffer", "Expiry"]
    for i, h in enumerate(acc_headers):
        c = ws.cell(row=11, column=2 + i, value=h)
        c.font = FONT_HDR; c.fill = FILL_HDR; c.alignment = ALIGN_C; c.border = BORDER_THIN

    def sc2(row, col, val, font=FONT_BODY, align=ALIGN_L, fill_=None):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = font; cell.alignment = align; cell.border = BORDER_THIN
        if fill_: cell.fill = fill_

    for j, a in enumerate(accum):
        r = 12 + j
        status = a.get("status", "SAFE").upper()
        sfont, fill = get_status_style(status)
        buf = a.get("buffer_pct", 0)

        sc2(r, 2, fmt_code(a.get("account", "")), FONT_BODY_B, ALIGN_C, fill)
        sc2(r, 3, a.get("description", ""), FONT_BODY, ALIGN_L, fill)
        sc2(r, 4, a.get("strike_call", ""), FONT_BODY, ALIGN_C, fill)
        sc2(r, 5, f"${a.get('current', 0):.2f}", FONT_BODY, ALIGN_R, fill)
        sc2(r, 6, f"{buf:+.1f}%", get_buffer_font(buf) if buf >= 0 else FONT_WARN, ALIGN_C, fill)
        sc2(r, 7, a.get("expiry", ""), FONT_BODY, ALIGN_C, fill)



def main():
    parser = argparse.ArgumentParser(description="Generate Portfolio XLSX Report")
    parser.add_argument("--output", "-o", default=DEFAULT_OUTPUT, help="Output .xlsx path")
    parser.add_argument("--json-file", help="Use pre-existing JSON instead of running portfolio_checker.py")
    args = parser.parse_args()
    
    # Get data
    if args.json_file:
        with open(args.json_file) as f:
            data = json.load(f)
    else:
        data = run_portfolio_checker()
    
    # Load template
    if not os.path.exists(TEMPLATE_PATH):
        print(f"ERROR: Template not found at {TEMPLATE_PATH}", file=sys.stderr)
        sys.exit(1)
    
    wb = load_workbook(TEMPLATE_PATH)
    
    # Load account display mapping

    # Populate each sheet
    populate_cover(wb["Cover"], data)
    populate_dashboard(wb["Dashboard"], data)
    populate_structured_notes(wb["Structured Notes"], data)
    populate_sell_puts(wb["Sell Puts"], data)
    populate_fx_accumulators(wb["FX & Accumulators"], data)
    

    # Save
    wb.save(args.output)
    print(f"OK: saved to {args.output}")
    print(f"Sheets: {wb.sheetnames}")
    print(f"Date: {date.today().isoformat()}")


if __name__ == "__main__":
    main()
